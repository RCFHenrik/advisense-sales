import csv
import io
import json
from datetime import date, datetime
from typing import Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy.orm import Session

from app.models.models import (
    Contact, Meeting, ColumnMapping, ContactStatusEnum, JobTitleDomain,
    ClassificationLookup, Employee, BusinessArea, Team, Site,
    LanguageEnum, RoleEnum, ApprovalStatusEnum, SiteLanguage, EmployeeSiteLanguage,
    ExpertiseTag,
)
from app.schemas.schemas import UploadDiffSummary, ConsultantUploadSummary


class ExcelImportService:
    def __init__(self, db: Session):
        self.db = db

    def _get_mappings(self, file_type: str) -> Dict[str, str]:
        """Get logical→physical column mappings for a file type."""
        mappings = self.db.query(ColumnMapping).filter(ColumnMapping.file_type == file_type).all()
        return {m.logical_field: m.physical_column for m in mappings}

    def _get_required_fields(self, file_type: str) -> List[str]:
        """Get list of required logical fields."""
        mappings = (
            self.db.query(ColumnMapping)
            .filter(ColumnMapping.file_type == file_type, ColumnMapping.is_required == True)
            .all()
        )
        return [m.logical_field for m in mappings]

    def _read_file(self, content: bytes, filename: str = "") -> Tuple[List[str], List[Dict]]:
        """Read Excel or CSV file and return headers + rows as list of dicts."""
        is_csv = filename.lower().endswith(".csv")

        if not is_csv:
            # Try to detect CSV by checking if content starts with text (BOM or ASCII)
            try:
                text_start = content[:4]
                if text_start.startswith(b'\xef\xbb\xbf') or all(
                    b < 128 for b in text_start
                ):
                    # Likely text/CSV — try to parse as CSV
                    is_csv = True
                else:
                    is_csv = False
            except Exception:
                is_csv = False

        if is_csv:
            return self._read_csv(content)
        else:
            return self._read_excel(content)

    def _read_csv(self, content: bytes) -> Tuple[List[str], List[Dict]]:
        """Read CSV file with auto-detected delimiter (semicolon or pipe)."""
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("cp1252")  # fallback for Windows-encoded files
        # Auto-detect delimiter: check first line for pipes vs semicolons
        first_line = text.split("\n")[0]
        if "|" in first_line and ";" not in first_line:
            delimiter = "|"
        else:
            delimiter = ";"  # existing default
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = reader.fieldnames or []
        data = []
        for row in reader:
            if any(v is not None and str(v).strip() for v in row.values()):
                data.append(row)
        if not data:
            raise ValueError("Empty file")
        return list(headers), data

    def _read_excel(self, content: bytes) -> Tuple[List[str], List[Dict]]:
        """Read Excel file and return headers + rows as list of dicts."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("Empty file")

        headers = [str(h).strip() if h else "" for h in rows[0]]
        data = []
        for row in rows[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = val
            if any(v is not None for v in row_dict.values()):
                data.append(row_dict)

        wb.close()
        return headers, data

    def _resolve_value(self, row: Dict, mappings: Dict[str, str], logical_field: str) -> Optional[str]:
        """Get value from row using the configured column mapping."""
        physical = mappings.get(logical_field)
        if not physical:
            return None
        val = row.get(physical)
        if val is None or str(val).strip() in ("", "#N/A", "N/A", "nan"):
            return None
        return str(val).strip()

    def _resolve_float(self, row: Dict, mappings: Dict[str, str], logical_field: str) -> Optional[float]:
        val = self._resolve_value(row, mappings, logical_field)
        if val is None:
            return None
        try:
            return float(val)
        except (ValueError, TypeError):
            return None

    def _resolve_int(self, row: Dict, mappings: Dict[str, str], logical_field: str) -> Optional[int]:
        val = self._resolve_value(row, mappings, logical_field)
        if val is None:
            return None
        try:
            return int(float(val))
        except (ValueError, TypeError):
            return None

    def _resolve_datetime(self, row: Dict, mappings: Dict[str, str], logical_field: str) -> Optional[datetime]:
        physical = mappings.get(logical_field)
        if not physical:
            return None
        val = row.get(physical)
        if val is None:
            return None
        if isinstance(val, datetime):
            return val
        val_str = str(val).strip()
        if not val_str or val_str in ("#N/A", "N/A", "nan"):
            return None
        try:
            return datetime.fromisoformat(val_str)
        except (ValueError, TypeError):
            pass
        # Try common date formats
        for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(val_str, fmt)
            except ValueError:
                continue
        return None

    def _resolve_bool(self, row: Dict, mappings: Dict[str, str], logical_field: str) -> bool:
        val = self._resolve_value(row, mappings, logical_field)
        if val is None:
            return False
        return val.lower() in ("true", "1", "yes", "x")

    def _resolve_optional_bool(self, row: Dict, mappings: Dict[str, str], logical_field: str):
        """Like _resolve_bool but returns None when the field is absent/empty."""
        val = self._resolve_value(row, mappings, logical_field)
        if val is None:
            return None
        return val.lower() in ("true", "1", "yes", "x")

    def _dedup_key(self, row: Dict, mappings: Dict[str, str]) -> Optional[str]:
        """Build deduplication key: email > record_id > (company_name, full_name)."""
        email = self._resolve_value(row, mappings, "email")
        if email:
            return f"email:{email.lower()}"

        record_id = self._resolve_value(row, mappings, "record_id")
        if record_id:
            return f"rid:{record_id}"

        company_name = self._resolve_value(row, mappings, "company_name")
        full_name = self._resolve_value(row, mappings, "full_name")
        if company_name and full_name:
            return f"comp:{company_name.lower()}:{full_name.lower()}"

        return None

    def import_contacts(self, content: bytes, batch_id: str, uploaded_by_id: int, filename: str = "") -> UploadDiffSummary:
        mappings = self._get_mappings("contacts")
        if not mappings:
            mappings = self._default_contact_mappings()
            self._save_default_mappings("contacts", mappings)

        required = self._get_required_fields("contacts")
        headers, data = self._read_file(content, filename)

        # Validate required mappings exist in headers
        errors = []
        for field in required:
            physical = mappings.get(field)
            if physical and physical not in headers:
                errors.append(f"Required column '{physical}' (for {field}) not found in file")

        if errors:
            raise ValueError("; ".join(errors))

        # Load job title → domain mappings for derivation
        domain_map = {}
        for jtd in self.db.query(JobTitleDomain).all():
            if jtd.domain and jtd.job_title:
                domain_map[jtd.job_title.lower()] = jtd.domain

        # Build lookup of existing contacts by dedup key
        existing_contacts = self.db.query(Contact).all()
        existing_map: Dict[str, Contact] = {}
        for c in existing_contacts:
            if c.email:
                existing_map[f"email:{c.email.lower()}"] = c
            elif c.record_id:
                existing_map[f"rid:{c.record_id}"] = c
            elif c.company_name and c.full_name:
                existing_map[f"comp:{c.company_name.lower()}:{c.full_name.lower()}"] = c

        added = 0
        updated = 0
        seen_keys = set()
        today = date.today()

        for row in data:
            key = self._dedup_key(row, mappings)
            if not key:
                continue
            if key in seen_keys:
                continue
            seen_keys.add(key)

            last_activity_date = self._resolve_datetime(row, mappings, "last_activity_date")

            # Calculate days_since_interaction and contacted_last_1y from LastContacted
            days_since = None
            contacted_1y = False
            if last_activity_date:
                days_since = (today - last_activity_date.date()).days
                if days_since < 0:
                    days_since = 0
                contacted_1y = days_since <= 365

            contact_data = {
                "record_id": self._resolve_value(row, mappings, "record_id"),
                "full_name": self._resolve_value(row, mappings, "full_name"),
                "email": self._resolve_value(row, mappings, "email"),
                "job_title": self._resolve_value(row, mappings, "job_title"),
                "company_name": self._resolve_value(row, mappings, "company_name"),
                "client_name": self._resolve_value(row, mappings, "client_name"),
                "sector": self._resolve_value(row, mappings, "sector"),
                "client_tier": self._resolve_value(row, mappings, "client_tier"),
                "group_domicile": self._resolve_value(row, mappings, "group_domicile"),
                "owner_name": self._resolve_value(row, mappings, "owner_name"),
                "owner_business_area": self._resolve_value(row, mappings, "owner_business_area"),
                "owner_org_site": self._resolve_value(row, mappings, "owner_org_site"),
                "owner_team": self._resolve_value(row, mappings, "owner_team"),
                "owner_seniority": self._resolve_value(row, mappings, "owner_seniority"),
                "last_activity_date": last_activity_date,
                "revenue": self._resolve_float(row, mappings, "revenue"),
                "has_historical_revenue": self._resolve_bool(row, mappings, "has_historical_revenue"),
                "days_since_interaction": days_since,
                "contacted_last_1y": contacted_1y,
                "hubspot_import_batch": batch_id,
            }

            # Normalize expert_areas to JSON array
            raw_ea = self._resolve_value(row, mappings, "expert_areas")
            if raw_ea:
                if raw_ea.strip().startswith("["):
                    contact_data["expert_areas"] = raw_ea.strip()
                else:
                    tags = [t.strip() for t in raw_ea.split(",") if t.strip()]
                    contact_data["expert_areas"] = json.dumps(tags) if tags else None

            # Normalize relevance_tags to JSON array
            raw_rt = self._resolve_value(row, mappings, "relevance_tags")
            if raw_rt:
                if raw_rt.strip().startswith("["):
                    contact_data["relevance_tags"] = raw_rt.strip()
                else:
                    rt_tags = [t.strip() for t in raw_rt.split(",") if t.strip()]
                    contact_data["relevance_tags"] = json.dumps(rt_tags) if rt_tags else None

            # Decision maker flag
            dm_val = self._resolve_value(row, mappings, "is_decision_maker")
            if dm_val is not None:
                contact_data["is_decision_maker"] = str(dm_val).strip().lower() in ("true", "1", "yes", "x")

            # Opt-out flags
            optout_1on1 = self._resolve_value(row, mappings, "opt_out_one_on_one")
            if optout_1on1 is not None:
                contact_data["opt_out_one_on_one"] = str(optout_1on1).strip().lower() in ("true", "1", "yes", "x")

            optout_marketing = self._resolve_value(row, mappings, "opt_out_marketing_info")
            if optout_marketing is not None:
                contact_data["opt_out_marketing_info"] = str(optout_marketing).strip().lower() in ("true", "1", "yes", "x")

            # Derive responsibility_domain from job title if available
            job_title = contact_data.get("job_title")
            if job_title and job_title.lower() in domain_map:
                contact_data["responsibility_domain"] = domain_map[job_title.lower()]

            # Also support direct mapping if present in file (backwards compat)
            if "responsibility_domain" in mappings and not contact_data.get("responsibility_domain"):
                direct_domain = self._resolve_value(row, mappings, "responsibility_domain")
                if direct_domain:
                    contact_data["responsibility_domain"] = direct_domain

            existing = existing_map.get(key)
            if existing:
                file_lad = contact_data.get('last_activity_date')
                _file_lad_accepted = False
                if file_lad is not None:
                    existing_lad = existing.last_activity_date
                    file_lad_n = file_lad.replace(tzinfo=None) if getattr(file_lad, 'tzinfo', None) else file_lad
                    existing_lad_n = existing_lad.replace(tzinfo=None) if existing_lad and getattr(existing_lad, 'tzinfo', None) else existing_lad
                    _file_lad_accepted = (existing_lad_n is None or file_lad_n > existing_lad_n)

                for k, v in contact_data.items():
                    if v is None:
                        continue
                    if k == 'last_activity_date':
                        if _file_lad_accepted:
                            existing.last_activity_date = v
                        continue
                    if k in ('days_since_interaction', 'contacted_last_1y'):
                        if _file_lad_accepted:
                            setattr(existing, k, v)
                        continue
                    # Handle job_title: preserve user edits, update original
                    if k == 'job_title':
                        existing.original_job_title = v  # always track latest CRM value
                        # If user has edited the title (job_title != original), keep their edit
                        if existing.original_job_title and existing.job_title != existing.original_job_title:
                            continue  # skip overwriting user's edited title
                    setattr(existing, k, v)
                existing.status = ContactStatusEnum.ACTIVE
                updated += 1
            else:
                # Set original_job_title from CRM import
                if contact_data.get("job_title"):
                    contact_data["original_job_title"] = contact_data["job_title"]
                contact = Contact(**contact_data)
                self.db.add(contact)
                added += 1

        # Mark contacts not in upload as inactive
        removed = 0
        for key, contact in existing_map.items():
            if key not in seen_keys and contact.status == ContactStatusEnum.ACTIVE:
                contact.status = ContactStatusEnum.INACTIVE_MISSING
                removed += 1

        self.db.commit()

        return UploadDiffSummary(
            added=added,
            updated=updated,
            removed=removed,
            unchanged=len(seen_keys) - added - updated,
            total_rows=len(data),
            errors=errors,
        )

    def import_meetings(self, content: bytes, batch_id: str, uploaded_by_id: int, filename: str = "") -> UploadDiffSummary:
        mappings = self._get_mappings("meetings")
        if not mappings:
            mappings = self._default_meeting_mappings()
            self._save_default_mappings("meetings", mappings)

        headers, data = self._read_file(content, filename)
        errors = []

        existing_meetings = self.db.query(Meeting).all()
        existing_by_rid = {m.record_id: m for m in existing_meetings if m.record_id}

        # Pre-load contact lookup (name → id) to avoid N+1 queries
        all_contacts = self.db.query(Contact.id, Contact.full_name).filter(
            Contact.full_name.isnot(None)
        ).all()
        contact_by_name: dict = {}
        for cid, cname in all_contacts:
            if cname:
                contact_by_name[cname.strip().lower()] = cid

        added = 0
        updated = 0
        seen_rids = set()
        batch_size = 500

        for i, row in enumerate(data):
            rid = self._resolve_value(row, mappings, "record_id")
            if rid:
                seen_rids.add(rid)

            meeting_data = {
                "record_id": rid,
                "employee_name": self._resolve_value(row, mappings, "employee_name"),
                "activity_date": self._resolve_datetime(row, mappings, "activity_date"),
                "details": self._resolve_value(row, mappings, "details"),
                "outcome": self._resolve_value(row, mappings, "outcome"),
                "associated_company": self._resolve_value(row, mappings, "associated_company"),
                "associated_contacts": self._resolve_value(row, mappings, "associated_contacts"),
                "client_tier": self._resolve_value(row, mappings, "client_tier"),
                "group_domicile": self._resolve_value(row, mappings, "group_domicile"),
                "seniority": self._resolve_value(row, mappings, "seniority"),
                "business_area": self._resolve_value(row, mappings, "business_area"),
                "team": self._resolve_value(row, mappings, "team"),
                "site": self._resolve_value(row, mappings, "site"),
                "client_name": self._resolve_value(row, mappings, "client_name"),
                "sector": self._resolve_value(row, mappings, "sector"),
                "hubspot_import_batch": batch_id,
            }

            # Try to link to contact via in-memory lookup
            contact_name = self._resolve_value(row, mappings, "associated_contacts")
            if contact_name:
                cid = contact_by_name.get(contact_name.strip().lower())
                if cid:
                    meeting_data["contact_id"] = cid

            existing = existing_by_rid.get(rid) if rid else None
            if existing:
                for k, v in meeting_data.items():
                    if v is not None:
                        setattr(existing, k, v)
                updated += 1
            else:
                meeting = Meeting(**meeting_data)
                self.db.add(meeting)
                added += 1

            # Flush in batches to avoid large memory buildup
            if (i + 1) % batch_size == 0:
                self.db.flush()

        self.db.commit()

        return UploadDiffSummary(
            added=added,
            updated=updated,
            removed=0,
            unchanged=len(seen_rids) - updated,
            total_rows=len(data),
            errors=errors,
        )

    def import_jobtitle_domain(self, content: bytes) -> UploadDiffSummary:
        """Import JobTitle→Domain mappings from Excel file."""
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        # Try to find the LLM_GROUPINGS sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "GROUPING" in name.upper() or "LLM" in name.upper():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        if not rows:
            raise ValueError("Empty file")

        # Build existing lookup
        existing = {jtd.job_title.lower(): jtd for jtd in self.db.query(JobTitleDomain).all()}

        added = 0
        updated = 0
        total = 0

        for row in rows[1:]:  # skip header
            vals = list(row)
            # Columns: (None, None, Num, JobTitle, Translation, LLM-GROUPINGS, PROPOSED AGGREGATES)
            if len(vals) < 6:
                continue

            job_title = str(vals[3]).strip() if vals[3] else ""
            domain = str(vals[5]).strip() if vals[5] else ""

            if not job_title:
                continue

            total += 1

            if domain == "":
                domain = "No Domain Classification"

            existing_jtd = existing.get(job_title.lower())
            if existing_jtd:
                if domain is not None and existing_jtd.domain != domain:
                    existing_jtd.domain = domain
                    updated += 1
            else:
                jtd = JobTitleDomain(job_title=job_title, domain=domain)
                self.db.add(jtd)
                added += 1

        self.db.commit()

        # Apply domain mappings to all contacts with matching job titles
        contacts_updated = self._apply_domains_to_contacts()

        summary = UploadDiffSummary(
            added=added,
            updated=updated,
            removed=0,
            unchanged=total - added - updated,
            total_rows=total,
            errors=[],
        )
        if contacts_updated > 0:
            summary.errors.append(f"✓ {contacts_updated} contacts updated with domain mappings")
        return summary

    def _apply_domains_to_contacts(self) -> int:
        """Update responsibility_domain on all contacts based on jobtitle_domains mappings."""
        # Build lowercase lookup: job_title → domain (include all mappings)
        jtd_map = {}
        for jtd in self.db.query(JobTitleDomain).all():
            domain = jtd.domain if jtd.domain else "No Domain Classification"
            jtd_map[jtd.job_title.strip().lower()] = domain

        if not jtd_map:
            return 0

        # Fetch contacts that have a job_title
        contacts = self.db.query(Contact).filter(Contact.job_title.isnot(None)).all()
        updated = 0
        for c in contacts:
            if not c.job_title:
                continue
            domain = jtd_map.get(c.job_title.strip().lower())
            if domain is not None and c.responsibility_domain != domain:
                c.responsibility_domain = domain
                updated += 1

        if updated:
            self.db.commit()
        return updated

    def import_classification(self, content: bytes, filename: str = "") -> UploadDiffSummary:
        """Import classification lookup data from CSV."""
        headers, data = self._read_file(content, filename)

        # Clear existing classification data (full replace)
        self.db.query(ClassificationLookup).delete()

        added = 0
        field_map = {
            "JobTitle": "job_title",
            "ClientGroupDomicile": "client_group_domicile",
            "ClientTier": "client_tier",
            "ClientIndustry": "client_industry",
            "NumContacts": "num_contacts",
            "Meetings_Total": "meetings_total",
            "TopBA_1": "top_ba_1",
            "TopBA_1_Count": "top_ba_1_count",
            "TopBA_1_Share": "top_ba_1_share",
            "TopBA_2": "top_ba_2",
            "TopBA_2_Count": "top_ba_2_count",
            "TopBA_2_Share": "top_ba_2_share",
            "TopTeam_1": "top_team_1",
            "TopTeam_1_Count": "top_team_1_count",
            "TopTeam_1_Share": "top_team_1_share",
            "TopTeam_2": "top_team_2",
            "TopTeam_2_Count": "top_team_2_count",
            "TopTeam_2_Share": "top_team_2_share",
            "TopRegistrator_1": "top_registrator_1",
            "TopRegistrator_1_Count": "top_registrator_1_count",
            "TopRegistrator_1_Share": "top_registrator_1_share",
            "TopRegistrator_2": "top_registrator_2",
            "TopRegistrator_2_Count": "top_registrator_2_count",
            "TopRegistrator_2_Share": "top_registrator_2_share",
            "Meetings_JRA_SA": "meetings_jra_sa",
            "Meetings_Manager": "meetings_manager",
            "Meetings_Senior_Manager": "meetings_senior_manager",
            "Meetings_Director": "meetings_director",
            "Meetings_Managing_Director": "meetings_managing_director",
            "Meetings_Other": "meetings_other",
        }

        int_fields = {
            "num_contacts", "meetings_total",
            "top_ba_1_count", "top_ba_2_count",
            "top_team_1_count", "top_team_2_count",
            "top_registrator_1_count", "top_registrator_2_count",
            "meetings_jra_sa", "meetings_manager", "meetings_senior_manager",
            "meetings_director", "meetings_managing_director", "meetings_other",
        }
        float_fields = {
            "top_ba_1_share", "top_ba_2_share",
            "top_team_1_share", "top_team_2_share",
            "top_registrator_1_share", "top_registrator_2_share",
        }

        for row in data:
            record = {}
            for csv_col, db_col in field_map.items():
                val = row.get(csv_col)
                if val is None or str(val).strip() in ("", "#N/A", "N/A", "nan"):
                    record[db_col] = None
                elif db_col in int_fields:
                    try:
                        record[db_col] = int(float(str(val).strip()))
                    except (ValueError, TypeError):
                        record[db_col] = None
                elif db_col in float_fields:
                    try:
                        # Handle comma decimal separator
                        record[db_col] = float(str(val).strip().replace(",", "."))
                    except (ValueError, TypeError):
                        record[db_col] = None
                else:
                    record[db_col] = str(val).strip()

            self.db.add(ClassificationLookup(**record))
            added += 1

        self.db.commit()

        return UploadDiffSummary(
            added=added,
            updated=0,
            removed=0,
            unchanged=0,
            total_rows=len(data),
            errors=[],
        )

    # ── Consultant batch import ─────────────────────────────────────────

    def _get_col(self, row: Dict, variants: List[str]) -> Optional[str]:
        """Try multiple column name variants (case-insensitive) and return the first non-empty match."""
        # Build a lowercase→value map once for case-insensitive matching
        row_lower = {str(k).strip().lower(): v for k, v in row.items()}
        for v in variants:
            val = row_lower.get(v.lower())
            if val is not None:
                s = str(val).strip()
                if s and s.lower() not in ("", "#n/a", "n/a", "nan", "none"):
                    return s
        return None

    def _is_active_consultant(self, row: Dict) -> bool:
        """Check if a consultant row represents a currently active employee.
        Returns True only if ValidToDT is 9999-12-31 (or the column is absent).
        """
        valid_to_raw = self._get_col(row, [
            "ValidToDT", "ValidToDate", "Valid To", "ValidTo",
            "EndDate", "End Date", "SlutDatum",
        ])
        if valid_to_raw is None:
            # Column not present — treat as active
            return True

        # Handle datetime objects (from Excel)
        if isinstance(valid_to_raw, (datetime, date)):
            d = valid_to_raw if isinstance(valid_to_raw, date) else valid_to_raw.date()
            return d.year == 9999

        # String comparison
        s = str(valid_to_raw).strip()
        return s.startswith("9999")

    def _parse_consultant_row(self, row: Dict, i: int, ba_map, team_map, site_map, lang_map, sl_by_name, warnings):
        """Parse a single consultant row, returning resolved fields dict or None for empty rows."""
        name = self._get_col(row, [
            "Name", "name", "FullName", "Full Name", "Namn",
            "ConsultantName", "Consultant Name", "Consultant",
        ])
        email = self._get_col(row, ["Email", "email", "E-mail", "E-post", "Epost"])
        seniority = self._get_col(row, [
            "Seniority", "seniority", "Senioritet",
            "ConsultantSeniority", "Consultant Seniority",
        ])
        ba_name = self._get_col(row, ["Business Area", "BusinessArea", "BA", "Affärsområde"])
        team_name = self._get_col(row, ["Team", "team"])
        site_name = self._get_col(row, ["Site", "site", "Office", "Kontor"])
        lang_raw = self._get_col(row, ["Language", "language", "Lang", "Språk"])
        site_lang_raw = self._get_col(row, [
            "SiteLanguage", "Site Language", "SiteLanguages", "Site Languages",
            "Språk (site)", "SiteSpråk",
        ])

        if not name and not email:
            return None

        # ── Resolve or auto-create BA ────────────────────────────────
        ba_id = None
        if ba_name:
            ba_id = ba_map.get(ba_name.lower())
            if not ba_id:
                new_ba = BusinessArea(name=ba_name)
                self.db.add(new_ba)
                self.db.flush()
                ba_map[ba_name.lower()] = new_ba.id
                ba_id = new_ba.id
                warnings.append(f"Row {i}: Business Area '{ba_name}' auto-created")

        # ── Resolve or auto-create Team (linked to BA if known) ───
        team_id = None
        if team_name:
            team_entry = team_map.get(team_name.lower())
            if team_entry:
                team_id = team_entry[0]
            elif ba_id is not None:
                # Team requires a business_area_id (NOT NULL FK)
                new_team = Team(name=team_name, business_area_id=ba_id)
                self.db.add(new_team)
                self.db.flush()
                team_map[team_name.lower()] = (new_team.id, ba_id)
                team_id = new_team.id
                warnings.append(f"Row {i}: Team '{team_name}' auto-created")
            else:
                warnings.append(f"Row {i}: Team '{team_name}' skipped (no Business Area)")

        # ── Resolve or auto-create Site ───────────────────────────
        site_id = None
        if site_name:
            site_id = site_map.get(site_name.lower())
            if not site_id:
                # Derive country_code from site name (e.g. "Site SE" → "SE")
                country_code = "XX"
                parts = site_name.split()
                if len(parts) >= 2 and parts[0].lower() == "site":
                    code_candidate = parts[-1].upper()
                    if 2 <= len(code_candidate) <= 5:
                        country_code = code_candidate
                new_site = Site(name=site_name, country_code=country_code)
                self.db.add(new_site)
                self.db.flush()
                site_map[site_name.lower()] = new_site.id
                site_id = new_site.id
                warnings.append(f"Row {i}: Site '{site_name}' auto-created (country_code={country_code})")

        language = LanguageEnum.ENGLISH
        if lang_raw:
            language = lang_map.get(lang_raw.lower(), LanguageEnum.ENGLISH)
        elif site_lang_raw:
            # Derive primary language from SiteLanguage when no Language column
            first_lang = site_lang_raw.replace(";", ",").split(",")[0].strip()
            if first_lang:
                language = lang_map.get(first_lang.lower(), LanguageEnum.ENGLISH)

        resolved_sl_ids: List[int] = []
        if site_lang_raw:
            parts = [p.strip() for p in site_lang_raw.replace(";", ",").split(",") if p.strip()]
            for part in parts:
                sl_id = sl_by_name.get(part.lower())
                if sl_id:
                    if sl_id not in resolved_sl_ids:
                        resolved_sl_ids.append(sl_id)
                else:
                    new_sl = SiteLanguage(name=part, is_active=True)
                    self.db.add(new_sl)
                    self.db.flush()
                    sl_by_name[part.lower()] = new_sl.id
                    resolved_sl_ids.append(new_sl.id)
                    warnings.append(f"Row {i}: SiteLanguage '{part}' auto-created")

        return {
            "name": name, "email": email, "seniority": seniority,
            "ba_id": ba_id, "team_id": team_id, "site_id": site_id,
            "language": language, "resolved_sl_ids": resolved_sl_ids,
        }

    def _build_lookup_maps(self):
        """Build common lookup maps used by consultant import."""
        ba_map = {ba.name.lower(): ba.id for ba in self.db.query(BusinessArea).all()}
        team_map: Dict[str, Tuple[int, int]] = {}
        for t in self.db.query(Team).all():
            team_map[t.name.lower()] = (t.id, t.business_area_id)
        site_map = {s.name.lower(): s.id for s in self.db.query(Site).all()}

        sl_by_name: Dict[str, int] = {}
        for sl in self.db.query(SiteLanguage).all():
            sl_by_name[sl.name.lower()] = sl.id
            if sl.code:
                sl_by_name[sl.code.lower()] = sl.id
            # Reactivate inactive site languages so they can be assigned
            if not sl.is_active:
                sl.is_active = True

        lang_map: Dict[str, LanguageEnum] = {
            "sv": LanguageEnum.SWEDISH, "swedish": LanguageEnum.SWEDISH, "svenska": LanguageEnum.SWEDISH,
            "no": LanguageEnum.NORWEGIAN, "norwegian": LanguageEnum.NORWEGIAN, "norsk": LanguageEnum.NORWEGIAN,
            "da": LanguageEnum.DANISH, "danish": LanguageEnum.DANISH, "dansk": LanguageEnum.DANISH,
            "en": LanguageEnum.ENGLISH, "english": LanguageEnum.ENGLISH, "engelska": LanguageEnum.ENGLISH,
            "de": LanguageEnum.GERMAN, "german": LanguageEnum.GERMAN, "tyska": LanguageEnum.GERMAN,
            "fi": LanguageEnum.FINNISH, "finnish": LanguageEnum.FINNISH, "finska": LanguageEnum.FINNISH,
        }

        return ba_map, team_map, site_map, sl_by_name, lang_map

    def import_consultants(
        self, content: bytes, batch_id: str, uploaded_by_id: int, filename: str = ""
    ) -> ConsultantUploadSummary:
        """Import consultants from CSV/Excel. All fields optional. Creates as PENDING."""
        headers, data = self._read_file(content, filename)

        ba_map, team_map, site_map, sl_by_name, lang_map = self._build_lookup_maps()
        existing_emails = set()
        existing_names = set()
        for e in self.db.query(Employee).all():
            if e.email and "@placeholder.local" not in e.email:
                existing_emails.add(e.email.lower())
            if e.name:
                existing_names.add(e.name.strip().lower())

        added = 0
        skipped = 0
        warnings: List[str] = []
        pending_site_langs: List[Tuple] = []

        for i, row in enumerate(data, start=2):
            parsed = self._parse_consultant_row(row, i, ba_map, team_map, site_map, lang_map, sl_by_name, warnings)
            if not parsed:
                continue

            email = parsed["email"]
            name = parsed["name"]

            # Skip if already exists (by email or by name)
            if email and email.lower() in existing_emails:
                skipped += 1
                continue
            if not email and name and name.strip().lower() in existing_names:
                skipped += 1
                continue

            emp_email = email or f"pending-{batch_id}-{i}@placeholder.local"

            emp = Employee(
                name=name or "(No name)",
                email=emp_email,
                role=RoleEnum.CONSULTANT,
                seniority=parsed["seniority"],
                business_area_id=parsed["ba_id"],
                team_id=parsed["team_id"],
                site_id=parsed["site_id"],
                primary_language=parsed["language"],
                is_active=False,
                approval_status=ApprovalStatusEnum.PENDING,
                uploaded_batch_id=batch_id,
            )
            self.db.add(emp)
            added += 1
            if email:
                existing_emails.add(email.lower())
            if name:
                existing_names.add(name.strip().lower())

            if parsed["resolved_sl_ids"]:
                pending_site_langs.append((emp, parsed["resolved_sl_ids"]))

        self.db.flush()
        for emp, sl_ids in pending_site_langs:
            for sl_id in sl_ids:
                esl = EmployeeSiteLanguage(employee_id=emp.id, site_language_id=sl_id)
                self.db.add(esl)

        self.db.commit()

        return ConsultantUploadSummary(
            added=added,
            skipped_duplicate=skipped,
            warnings=warnings,
            total_rows=len(data),
        )

    def import_consultants_upsert(
        self, content: bytes, batch_id: str, uploaded_by_id: int, filename: str = ""
    ) -> ConsultantUploadSummary:
        """Import consultants with upsert: add new as APPROVED, update existing fields."""
        headers, data = self._read_file(content, filename)

        ba_map, team_map, site_map, sl_by_name, lang_map = self._build_lookup_maps()

        # Build lookup of existing employees by email AND by name (fallback)
        existing_by_email: Dict[str, Employee] = {}
        existing_by_name: Dict[str, Employee] = {}
        for emp in self.db.query(Employee).all():
            if emp.email and "@placeholder.local" not in emp.email:
                existing_by_email[emp.email.lower()] = emp
            if emp.name:
                existing_by_name[emp.name.strip().lower()] = emp

        added = 0
        updated = 0
        deactivated = 0
        warnings: List[str] = []
        pending_site_langs: List[Tuple] = []

        for i, row in enumerate(data, start=2):
            # Check if consultant is still active (ValidToDT = 9999-12-31)
            is_active = self._is_active_consultant(row)

            parsed = self._parse_consultant_row(row, i, ba_map, team_map, site_map, lang_map, sl_by_name, warnings)
            if not parsed:
                continue

            email = parsed["email"]
            name = parsed["name"]

            # Check if this consultant already exists (by email first, then by name)
            existing_emp = None
            if email:
                existing_emp = existing_by_email.get(email.lower())
            if not existing_emp and name:
                existing_emp = existing_by_name.get(name.strip().lower())

            # If consultant has left (ValidToDT is a real date), deactivate if they exist
            if not is_active:
                if existing_emp and existing_emp.is_active:
                    existing_emp.is_active = False
                    deactivated += 1
                continue

            if existing_emp:
                # Upsert: update fields from file (file takes precedence)
                changed = False
                if parsed["name"] and parsed["name"] != existing_emp.name:
                    existing_emp.name = parsed["name"]
                    changed = True
                if parsed["seniority"] is not None and parsed["seniority"] != existing_emp.seniority:
                    existing_emp.seniority = parsed["seniority"]
                    changed = True
                if parsed["ba_id"] is not None and parsed["ba_id"] != existing_emp.business_area_id:
                    existing_emp.business_area_id = parsed["ba_id"]
                    changed = True
                if parsed["team_id"] is not None and parsed["team_id"] != existing_emp.team_id:
                    existing_emp.team_id = parsed["team_id"]
                    changed = True
                if parsed["site_id"] is not None and parsed["site_id"] != existing_emp.site_id:
                    existing_emp.site_id = parsed["site_id"]
                    changed = True
                if parsed["language"] != existing_emp.primary_language:
                    existing_emp.primary_language = parsed["language"]
                    changed = True

                # Update site languages: replace with file's set
                if parsed["resolved_sl_ids"]:
                    # Remove existing site language links
                    self.db.query(EmployeeSiteLanguage).filter(
                        EmployeeSiteLanguage.employee_id == existing_emp.id
                    ).delete()
                    for sl_id in parsed["resolved_sl_ids"]:
                        esl = EmployeeSiteLanguage(employee_id=existing_emp.id, site_language_id=sl_id)
                        self.db.add(esl)
                    changed = True

                # Ensure active + approved
                if not existing_emp.is_active:
                    existing_emp.is_active = True
                    existing_emp.approval_status = ApprovalStatusEnum.APPROVED
                    changed = True

                if changed:
                    updated += 1
            else:
                # New consultant — add as approved
                emp_email = email or f"batch-{batch_id}-{i}@placeholder.local"

                emp = Employee(
                    name=parsed["name"] or "(No name)",
                    email=emp_email,
                    role=RoleEnum.CONSULTANT,
                    seniority=parsed["seniority"],
                    business_area_id=parsed["ba_id"],
                    team_id=parsed["team_id"],
                    site_id=parsed["site_id"],
                    primary_language=parsed["language"],
                    is_active=True,
                    approval_status=ApprovalStatusEnum.APPROVED,
                    uploaded_batch_id=batch_id,
                )
                self.db.add(emp)
                added += 1
                if email:
                    existing_by_email[email.lower()] = emp
                if name:
                    existing_by_name[name.strip().lower()] = emp

                if parsed["resolved_sl_ids"]:
                    pending_site_langs.append((emp, parsed["resolved_sl_ids"]))

        # Flush for new employee IDs, then create site language links
        self.db.flush()
        for emp, sl_ids in pending_site_langs:
            for sl_id in sl_ids:
                esl = EmployeeSiteLanguage(employee_id=emp.id, site_language_id=sl_id)
                self.db.add(esl)

        self.db.commit()

        if deactivated > 0:
            warnings.insert(0, f"{deactivated} former employees deactivated (ValidToDT is not 9999-12-31)")

        return ConsultantUploadSummary(
            added=added,
            updated=updated,
            warnings=warnings,
            total_rows=len(data),
        )

    def _default_contact_mappings(self) -> Dict[str, str]:
        return {
            "record_id": "RecordID",
            "email": "ContactEmail",
            "full_name": "ContactFullName",
            "job_title": "JobTitle",
            "company_name": "ClientGroup",
            "client_name": "ClientName",
            "sector": "ClientIndustry",
            "client_tier": "ClientTier",
            "group_domicile": "ClientGroupDomicile",
            "owner_name": "ContactOwner",
            "owner_business_area": "BusinessArea",
            "owner_org_site": "Site",
            "owner_team": "Team",
            "owner_seniority": "ContactOwnerSeniority",
            "last_activity_date": "LastContacted",
            "revenue": "TotalRevenue",
            "has_historical_revenue": "HasHistoricalRevenue",
            "expert_areas": "ExpertAreas",
            "relevance_tags": "RelevanceTags",
            "is_decision_maker": "IsDecisionMaker",
            "opt_out_one_on_one": "OptOut_OneOnOne",
            "opt_out_marketing_info": "OptOut_MarketingInfo",
        }

    def _default_meeting_mappings(self) -> Dict[str, str]:
        return {
            "record_id": "RecordID",
            "details": "MeetingDetails",
            "activity_date": "MeetingDate",
            "employee_name": "MeetingRegistrator",
            "seniority": "MeetingRegistratorSeniority",
            "business_area": "BusinessArea",
            "team": "Team",
            "site": "Site",
            "associated_company": "ClientGroup",
            "client_name": "ClientName",
            "client_tier": "ClientTier",
            "group_domicile": "ClientGroupDomicile",
            "sector": "ClientIndustry",
            "associated_contacts": "MeetingParticipant",
            "outcome": "MeetingRegistrationStatus",
        }

    def _save_default_mappings(self, file_type: str, mappings: Dict[str, str]):
        required_contacts = {"record_id", "email"}
        required_meetings = {"record_id", "activity_date"}

        required_set = required_contacts if file_type == "contacts" else required_meetings

        for logical, physical in mappings.items():
            m = ColumnMapping(
                file_type=file_type,
                logical_field=logical,
                physical_column=physical,
                is_required=logical in required_set,
            )
            self.db.add(m)
        self.db.commit()

    def import_expertise_tags(self, content: bytes, filename: str = "") -> UploadDiffSummary:
        """Import expertise/relevance tags from CSV/Excel.

        Supports two formats:
        1. Per-contact mapping: columns include ContactEmail + RelevanceTags
           -> updates Contact.relevance_tags for each matched contact
        2. Simple tag list: single column with tag names
           -> only updates the ExpertiseTag reference table

        Both formats also maintain the ExpertiseTag reference table.
        """
        headers, data = self._read_file(content, filename)

        if not headers:
            raise ValueError("File has no columns")

        # Detect per-contact mapping format
        headers_lower = {h.lower().strip(): h for h in headers}
        has_email_col = any(k in headers_lower for k in ("contactemail", "email", "contact_email"))
        has_tags_col = any(k in headers_lower for k in ("relevancetags", "relevance_tags", "tags"))

        if has_email_col and has_tags_col:
            return self._import_relevance_tags_per_contact(headers_lower, data)
        else:
            return self._import_expertise_tags_simple(headers, data)

    def _import_expertise_tags_simple(self, headers: List[str], data: List[Dict]) -> UploadDiffSummary:
        """Original simple import: one tag per row or comma-separated in first column."""
        first_col = headers[0]

        raw_tags = []
        for row in data:
            val = row.get(first_col)
            if val is None:
                continue
            cell = str(val).strip()
            if not cell:
                continue
            for tag in cell.split(","):
                tag = tag.strip()
                if tag:
                    raw_tags.append(tag)

        if not raw_tags:
            raise ValueError("No tags found in file")

        seen = {}
        for tag in raw_tags:
            key = tag.lower()
            if key not in seen:
                seen[key] = tag

        existing = {
            et.name.lower(): et
            for et in self.db.query(ExpertiseTag).all()
        }

        added = 0
        unchanged = 0
        reactivated = 0

        for key, tag_name in seen.items():
            if key in existing:
                et = existing[key]
                if not et.is_active:
                    et.is_active = True
                    reactivated += 1
                else:
                    unchanged += 1
            else:
                self.db.add(ExpertiseTag(name=tag_name, is_active=True))
                added += 1

        self.db.flush()

        return UploadDiffSummary(
            added=added,
            updated=reactivated,
            removed=0,
            unchanged=unchanged,
            total_rows=len(raw_tags),
            errors=[],
        )

    def _import_relevance_tags_per_contact(self, headers_lower: Dict[str, str], data: List[Dict]) -> UploadDiffSummary:
        """Import per-contact relevance tags: match by email, update Contact.relevance_tags."""
        from sqlalchemy import func

        # Resolve actual column names from headers
        email_col = None
        for key in ("contactemail", "email", "contact_email"):
            if key in headers_lower:
                email_col = headers_lower[key]
                break

        tags_col = None
        for key in ("relevancetags", "relevance_tags", "tags"):
            if key in headers_lower:
                tags_col = headers_lower[key]
                break

        if not email_col or not tags_col:
            raise ValueError("Could not find email and tags columns")

        # Build email -> tags mapping from file
        email_to_tags: Dict[str, List[str]] = {}
        all_tags_set: Dict[str, str] = {}  # lowercase -> original casing

        for row in data:
            email_val = row.get(email_col)
            tags_val = row.get(tags_col)
            if not email_val or not tags_val:
                continue
            email = str(email_val).strip().lower()
            if not email or "@" not in email:
                continue

            tags_str = str(tags_val).strip()
            if not tags_str:
                continue

            tags = [t.strip() for t in tags_str.split(",") if t.strip()]
            if not tags:
                continue

            email_to_tags[email] = tags
            for t in tags:
                key = t.lower()
                if key not in all_tags_set:
                    all_tags_set[key] = t

        if not email_to_tags:
            raise ValueError("No valid contact-tag mappings found in file")

        # 1. Update ExpertiseTag reference table
        existing_et = {
            et.name.lower(): et
            for et in self.db.query(ExpertiseTag).all()
        }

        et_added = 0
        et_reactivated = 0
        for key, tag_name in all_tags_set.items():
            if key in existing_et:
                et = existing_et[key]
                if not et.is_active:
                    et.is_active = True
                    et_reactivated += 1
            else:
                self.db.add(ExpertiseTag(name=tag_name, is_active=True))
                et_added += 1

        self.db.flush()

        # 2. Update Contact.relevance_tags for matched contacts
        # Load ALL contacts with email and build two lookup dicts:
        #   - exact email match
        #   - email prefix match (part before @) as fallback for anonymized data
        all_contacts = (
            self.db.query(Contact)
            .filter(Contact.email.isnot(None), Contact.email != "")
            .all()
        )

        contacts_by_email: Dict[str, Contact] = {}
        contacts_by_prefix: Dict[str, Contact] = {}
        for c in all_contacts:
            if c.email:
                email_lower = c.email.lower().strip()
                contacts_by_email[email_lower] = c
                prefix = email_lower.split("@")[0]
                contacts_by_prefix[prefix] = c

        contacts_updated = 0
        contacts_unchanged = 0
        contacts_not_found = 0
        matched_by_prefix = 0
        errors = []

        for email, tags in email_to_tags.items():
            # Try exact email match first
            contact = contacts_by_email.get(email)
            if not contact:
                # Fallback: match by email prefix (part before @)
                prefix = email.split("@")[0]
                contact = contacts_by_prefix.get(prefix)
                if contact:
                    matched_by_prefix += 1
            if not contact:
                contacts_not_found += 1
                continue

            new_tags_json = json.dumps(tags)
            if contact.relevance_tags == new_tags_json:
                contacts_unchanged += 1
            else:
                contact.relevance_tags = new_tags_json
                contacts_updated += 1

        self.db.flush()

        if contacts_not_found > 0:
            errors.append(f"{contacts_not_found} contacts not found in CRM (email not matched)")
        if matched_by_prefix > 0:
            errors.append(f"{matched_by_prefix} contacts matched by email prefix (domain differed)")

        return UploadDiffSummary(
            added=et_added,
            updated=contacts_updated,
            removed=0,
            unchanged=contacts_unchanged,
            total_rows=len(email_to_tags),
            errors=errors,
        )


    def import_coverage_gaps(self, content: bytes, batch_id: str, filename: str = "") -> dict:
        """Import coverage gap analysis from pipe-delimited CSV.

        The file has 3-4 description/header rows before actual data.
        Skip rows where 'Company' column is empty or looks like a description.
        """
        import json as _json
        from app.models.models import CoverageGap

        headers, data = self._read_file(content, filename)

        # Full replace
        self.db.query(CoverageGap).delete()

        added = 0
        skipped = 0
        total_critical = 0
        total_potential = 0

        def safe_int(val):
            if val is None:
                return 0
            try:
                return int(float(str(val).strip()))
            except (ValueError, TypeError):
                return 0

        def parse_list(val):
            """Parse comma-separated string to JSON array, skip dashes."""
            if not val:
                return None
            s = str(val).strip()
            if not s or s in ("\u2013", "\u2014", "-", "\u2013", "N/A", "nan", "None"):
                return None
            items = [v.strip() for v in s.split(",") if v.strip() and v.strip() not in ("\u2013", "\u2014", "-")]
            return _json.dumps(items) if items else None

        # Check if this looks like a description row
        skip_prefixes = ("industry peers", "=50%", "15-49%", "broad capability",
                         "senior roles", "capabilities that", "missing domains",
                         "missing titles", "company", "")

        for row in data:
            company = str(row.get("Company", "") or "").strip()

            # Skip description/empty rows
            if not company or company.lower() in skip_prefixes or len(company) < 3:
                skipped += 1
                continue

            # Skip rows that look like column descriptions (contain keywords)
            first_chars = company.lower()[:20]
            if any(kw in first_chars for kw in ["industry peer", "=50%", "15-49%", "broad cap", "senior role"]):
                skipped += 1
                continue

            gap = CoverageGap(
                company_name=company,
                company_name_normalized=company.lower().strip(),
                industry=str(row.get("Industry", "") or "").strip() or None,
                tier=str(row.get("Tier", "") or "").strip() or None,
                contacts_in_crm=safe_int(row.get("Contacts in CRM")),
                critical_gap_count=safe_int(row.get("Critical Gaps")),
                potential_gap_count=safe_int(row.get("Potential Gaps")),
                total_gap_count=safe_int(row.get("Total Gaps")),
                missing_domains_critical=parse_list(row.get("Missing Domains (most peers have this)")),
                missing_titles_critical=parse_list(row.get("Missing Titles (most peers have this)")),
                missing_domains_potential=parse_list(row.get("Missing Domains (some peers have this)")),
                missing_titles_potential=parse_list(row.get("Missing Titles (some peers have this)")),
                upload_batch_id=batch_id,
            )
            self.db.add(gap)
            added += 1
            total_critical += gap.critical_gap_count
            total_potential += gap.potential_gap_count

        self.db.commit()

        return {
            "added": added,
            "updated": 0,
            "removed": 0,
            "total_rows": len(data),
            "skipped_header_rows": skipped,
            "companies_with_gaps": added,
            "total_critical_gaps": total_critical,
            "total_potential_gaps": total_potential,
        }

